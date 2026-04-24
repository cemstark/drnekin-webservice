import argparse
import json
import logging
import mimetypes
import os
import sys
import time
import uuid
import re
from datetime import datetime
from pathlib import Path
from urllib import request, error


def load_env(path: Path) -> dict:
    values = {}
    if not path.exists():
        return values
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"')
    return values


def wait_until_readable(path: Path, attempts: int = 12, delay: float = 2.0) -> bool:
    last_size = -1
    stable_count = 0
    for _ in range(attempts):
        try:
            size = path.stat().st_size
            with path.open("rb") as handle:
                handle.read(1)
            if size == last_size:
                stable_count += 1
            else:
                stable_count = 0
            if stable_count >= 1:
                return True
            last_size = size
        except OSError:
            pass
        time.sleep(delay)
    return False


def multipart_body(field_name: str, file_path: Path) -> tuple[bytes, str]:
    boundary = "----drn-sync-" + uuid.uuid4().hex
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    parts = [
        f"--{boundary}\r\n".encode(),
        f'Content-Disposition: form-data; name="{field_name}"; filename="{file_path.name}"\r\n'.encode(),
        f"Content-Type: {content_type}\r\n\r\n".encode(),
        file_path.read_bytes(),
        b"\r\n",
        f"--{boundary}--\r\n".encode(),
    ]
    return b"".join(parts), boundary


def upload_excel(api_url: str, api_key: str, excel_path: Path) -> dict:
    body, boundary = multipart_body("excel", excel_path)
    req = request.Request(
        api_url,
        data=body,
        method="POST",
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Content-Length": str(len(body)),
            "X-Panel-Api-Key": api_key,
            "User-Agent": "DRN Excel Sync Agent/1.0",
        },
    )
    try:
        with request.urlopen(req, timeout=60) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw)
    except error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {"ok": False, "error": raw or str(exc)}


def api_json(url: str, api_key: str, payload: dict | None = None) -> dict:
    data = None
    method = "GET"
    headers = {
        "X-Panel-Api-Key": api_key,
        "User-Agent": "DRN Excel Sync Agent/1.0",
    }
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        method = "POST"
        headers["Content-Type"] = "application/json"

    req = request.Request(url, data=data, method=method, headers=headers)
    try:
        with request.urlopen(req, timeout=60) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw)
    except error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {"ok": False, "error": raw or str(exc)}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def header_key(value) -> str:
    text = "" if value is None else str(value).strip()
    table = str.maketrans({
        "İ": "i", "I": "i", "ı": "i", "Ğ": "g", "ğ": "g",
        "Ü": "u", "ü": "u", "Ş": "s", "ş": "s", "Ö": "o",
        "ö": "o", "Ç": "c", "ç": "c",
    })
    text = text.translate(table).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def date_value(value) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip().replace("O", "0").replace("o", "0")
    text = re.sub(r"\.+", ".", text)
    match = re.match(r"^(\d{2})(\d{2})(\d{4})$", text)
    if match:
        return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
    match = re.match(r"^(\d{1,2})\.(\d{1,2})\.(206)$", text)
    if match:
        return f"2026-{int(match.group(2)):02d}-{int(match.group(1)):02d}"
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def build_header_map(headers: list) -> dict:
    aliases = {
        "plate": ["plaka", "arac plakasi", "arac plaka"],
        "customer_name": ["ad soyad", "adi soyadi", "isim soyisim", "kullanici isim soyisim", "musteri", "musteri ad soyad"],
        "insurance_company": ["sigorta sirketi", "sigorta", "si gorta"],
        "repair_status": ["tamir durumu", "guncel durum", "servisteki guncel durum", "durum"],
        "mini_repair_part": ["mini onarim parca", "hangi parca", "parca", "prc"],
        "service_entry_date": ["giris tarihi", "servise giris tarihi", "ser giris", "ser gi ri"],
        "service_exit_date": ["cikis tarihi", "servisten cikis tarihi", "tes tarihi", "tes tari hi"],
        "file_no": ["dosya no", "dosya numarasi"],
    }
    normalized = [header_key(h) for h in headers]
    result = {}
    for index, header in enumerate(normalized):
        for field, names in aliases.items():
            if header in [header_key(name) for name in names]:
                result[field] = index + 1
    return result


def generated_record_no(sheet_name: str, plate: str, entry_date: str | None, file_no: str, line: int) -> str:
    parts = [
        header_key(sheet_name),
        header_key(plate),
        entry_date or "tarihsiz",
        header_key(file_no) if file_no else f"satir-{line}",
    ]
    return "-".join([part for part in parts if part])[:80]


def set_excel_cell(ws, row: int, column: int | None, value) -> None:
    if not column:
        return
    ws.cell(row=row, column=column).value = value


def workbook_is_open_in_excel(excel_path: Path):
    try:
        import win32com.client
    except ImportError:
        return None

    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return None

    wanted = str(excel_path.resolve()).lower()
    for workbook in excel.Workbooks:
        try:
            if str(workbook.FullName).lower() == wanted:
                return workbook
        except Exception:
            continue
    return None


def com_cell_value(ws, row: int, column: int):
    value = ws.Cells(row, column).Value
    return "" if value is None else value


def com_set_cell(ws, row: int, column: int | None, value) -> None:
    if not column:
        return
    ws.Cells(row, column).Value = "" if value is None else value


def apply_updates_to_open_workbook(excel_path: Path, updates: list) -> tuple[list, list] | None:
    workbook = workbook_is_open_in_excel(excel_path)
    if workbook is None:
        return None

    row_index = {}
    for sheet_number in range(1, workbook.Worksheets.Count + 1):
        ws = workbook.Worksheets(sheet_number)
        headers = [com_cell_value(ws, 1, col) for col in range(1, ws.UsedRange.Columns.Count + 1)]
        mapping = build_header_map(headers)
        if not {"plate", "customer_name", "service_entry_date"}.issubset(mapping):
            continue

        last_row = int(ws.Cells(ws.Rows.Count, mapping["plate"]).End(-4162).Row)
        for row in range(2, last_row + 1):
            values = [com_cell_value(ws, row, col) for col in range(1, ws.UsedRange.Columns.Count + 1)]
            if not any(value is not None and str(value).strip() for value in values):
                continue

            plate = str(com_cell_value(ws, row, mapping["plate"]) or "").strip()
            entry = date_value(com_cell_value(ws, row, mapping["service_entry_date"]))
            file_no = str(com_cell_value(ws, row, mapping.get("file_no", 0)) or "").strip() if mapping.get("file_no") else ""
            for sheet_name in (ws.Name, f"Sheet{sheet_number}"):
                record_no = generated_record_no(sheet_name, plate, entry, file_no, row)
                row_index[record_no] = (ws, row, mapping)

    applied = []
    failed = []
    for update in updates:
        update_id = int(update.get("id", 0))
        record_no = str(update.get("record_no", ""))
        fields = update.get("fields") or {}
        target = row_index.get(record_no)
        if not target:
            failed.append({"id": update_id, "error": f"Excel row not found for {record_no}"})
            continue

        ws, row, mapping = target
        try:
            com_set_cell(ws, row, mapping.get("plate"), fields.get("plate"))
            com_set_cell(ws, row, mapping.get("customer_name"), fields.get("customer_name"))
            com_set_cell(ws, row, mapping.get("insurance_company"), fields.get("insurance_company"))
            com_set_cell(ws, row, mapping.get("repair_status"), fields.get("repair_status"))
            com_set_cell(ws, row, mapping.get("mini_repair_part"), fields.get("mini_repair_part") if fields.get("mini_repair_has") else "")
            com_set_cell(ws, row, mapping.get("service_entry_date"), fields.get("service_entry_date"))
            com_set_cell(ws, row, mapping.get("service_exit_date"), fields.get("service_exit_date") or "")
            applied.append(update_id)
        except Exception as exc:
            failed.append({"id": update_id, "error": str(exc)})

    if applied:
        workbook.Save()
        logging.info("Applied %s panel update(s) to open Excel workbook", len(applied))

    return applied, failed


def apply_pending_updates(excel_path: Path, updates_url: str, api_key: str) -> int:
    pending = api_json(updates_url, api_key)
    if not pending.get("ok"):
        logging.error("Could not fetch panel updates: %s", pending.get("error") or pending)
        return 0

    updates = pending.get("updates") or []
    if not updates:
        return 0

    com_result = apply_updates_to_open_workbook(excel_path, updates)
    if com_result is not None:
        applied, failed = com_result
        result = api_json(updates_url, api_key, {"applied": applied, "failed": failed})
        if not result.get("ok"):
            logging.error("Could not confirm panel updates: %s", result.get("error") or result)
        for failure in failed:
            logging.error("Panel update failed: %s", failure)
        return len(applied)

    try:
        from openpyxl import load_workbook
    except ImportError:
        logging.error("openpyxl is missing. Run: python -m pip install openpyxl")
        return 0

    if not wait_until_readable(excel_path):
        logging.error("Excel file is locked; panel updates will retry later")
        return 0

    wb = load_workbook(excel_path)
    row_index = {}
    maps = {}
    for sheet_number, ws in enumerate(wb.worksheets, start=1):
        headers = [cell.value for cell in ws[1]]
        mapping = build_header_map(headers)
        maps[ws.title] = mapping
        if not {"plate", "customer_name", "service_entry_date"}.issubset(mapping):
            continue
        for row in range(2, ws.max_row + 1):
            values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
            if not any(value is not None and str(value).strip() for value in values):
                continue
            plate = str(ws.cell(row=row, column=mapping["plate"]).value or "").strip()
            entry = date_value(ws.cell(row=row, column=mapping["service_entry_date"]).value)
            file_no = str(ws.cell(row=row, column=mapping.get("file_no", 0)).value or "").strip() if mapping.get("file_no") else ""
            for sheet_name in (ws.title, f"Sheet{sheet_number}"):
                record_no = generated_record_no(sheet_name, plate, entry, file_no, row)
                row_index[record_no] = (ws, row, mapping)

    applied = []
    failed = []
    for update in updates:
        update_id = int(update.get("id", 0))
        record_no = str(update.get("record_no", ""))
        fields = update.get("fields") or {}
        target = row_index.get(record_no)
        if not target:
            failed.append({"id": update_id, "error": f"Excel row not found for {record_no}"})
            continue

        ws, row, mapping = target
        try:
            set_excel_cell(ws, row, mapping.get("plate"), fields.get("plate"))
            set_excel_cell(ws, row, mapping.get("customer_name"), fields.get("customer_name"))
            set_excel_cell(ws, row, mapping.get("insurance_company"), fields.get("insurance_company"))
            set_excel_cell(ws, row, mapping.get("repair_status"), fields.get("repair_status"))
            if fields.get("mini_repair_has"):
                set_excel_cell(ws, row, mapping.get("mini_repair_part"), fields.get("mini_repair_part"))
            else:
                set_excel_cell(ws, row, mapping.get("mini_repair_part"), "")
            set_excel_cell(ws, row, mapping.get("service_entry_date"), fields.get("service_entry_date"))
            set_excel_cell(ws, row, mapping.get("service_exit_date"), fields.get("service_exit_date") or "")
            applied.append(update_id)
        except Exception as exc:
            failed.append({"id": update_id, "error": str(exc)})

    if applied:
        wb.save(excel_path)
        logging.info("Applied %s panel update(s) to Excel", len(applied))

    result = api_json(updates_url, api_key, {"applied": applied, "failed": failed})
    if not result.get("ok"):
        logging.error("Could not confirm panel updates: %s", result.get("error") or result)

    for failure in failed:
        logging.error("Panel update failed: %s", failure)

    return len(applied)


def run_watch(excel_path: Path, api_url: str, api_key: str, interval: float, updates_url: str) -> None:
    last_mtime = None
    logging.info("Watching %s", excel_path)
    while True:
        if not excel_path.exists():
            logging.warning("Excel file not found: %s", excel_path)
            time.sleep(interval)
            continue

        applied_count = apply_pending_updates(excel_path, updates_url, api_key)
        if applied_count:
            sync_once(excel_path, api_url, api_key)
            last_mtime = excel_path.stat().st_mtime

        mtime = excel_path.stat().st_mtime
        if last_mtime is None:
            last_mtime = mtime
            logging.info("Initial upload")
            sync_once(excel_path, api_url, api_key)
        elif mtime != last_mtime:
            last_mtime = mtime
            logging.info("Change detected")
            sync_once(excel_path, api_url, api_key)

        time.sleep(interval)


def sync_once(excel_path: Path, api_url: str, api_key: str) -> None:
    if not wait_until_readable(excel_path):
        logging.error("Excel file is still locked or unstable: %s", excel_path)
        return
    result = upload_excel(api_url, api_key, excel_path)
    if result.get("ok"):
        payload = result.get("result", {})
        logging.info("Sync ok: imported=%s skipped=%s status=%s", payload.get("imported"), payload.get("skipped"), payload.get("status"))
    else:
        logging.error("Sync failed: %s", result.get("error") or result)


def main() -> int:
    parser = argparse.ArgumentParser(description="DRN Excel sync agent")
    parser.add_argument("--once", action="store_true", help="Run one upload and exit")
    parser.add_argument("--env", default=".env", help="Path to env file")
    args = parser.parse_args()

    env = load_env(Path(args.env))
    excel_path = Path(env.get("EXCEL_PATH", "")).expanduser()
    api_url = env.get("API_URL", "")
    updates_url = env.get("UPDATES_URL", api_url.replace("/api/import.php", "/api/excel-updates.php"))
    api_key = env.get("API_KEY", "")
    interval = float(env.get("POLL_SECONDS", "5"))
    log_path = Path(env.get("LOG_PATH", "sync-agent.log"))

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[logging.FileHandler(log_path, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )

    if not excel_path or not api_url or not api_key:
        logging.error("EXCEL_PATH, API_URL and API_KEY are required in .env")
        return 2

    if args.once:
        sync_once(excel_path, api_url, api_key)
    else:
        run_watch(excel_path, api_url, api_key, interval, updates_url)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
